from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

ARQUIVO = 'cardapio.xlsx'

if not os.path.exists(ARQUIVO):
    wb = Workbook()
    ws = wb.active
    ws.append([
        'Imagem', 'Produto', 'Ingredientes', 'Preço de custo', 'Margem', 'Preço de venda'
    ])
    wb.save(ARQUIVO)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/salvar', methods=["POST"])
def salvar():
    imagem = request.form['imagem']
    nome = request.form['nome']
    ingrediente = float(request.form['prova1'])
    custo = float(request.form['prova2'])
    margem = float(request.form['atividade'])
    venda = round(((prova1 + prova2 + atividade) / 3), 2)
    if media >= 6:
        situacao = 'APROVADO'
    elif media < 5:
        situacao = 'REPROVADO'
    else:
        situacao = 'RECUPERAÇÃO'
    workbook = load_workbook(ARQUIVO)
    ws_local = workbook.active
    ra = ws_local['H1'].value
    ws_local.append([nome, ra, prova1, prova2, atividade, media, situacao])
    ra += 1
    ws_local['H1'] = ra
    workbook.save(ARQUIVO)
    return render_template(
        'resultado.html',
        nome=nome,
        ra=ra,
        prova1=prova1,
        prova2=prova2,
        atividade=atividade,
        media=media,
        situacao=situacao
    )


# '@app.route('/analisar')
# def analisar():
#     nome_param = request.args.get('nome')
#     wb = load_workbook(ARQUIVO)
#     ws = wb.active
#     for linha in ws.iter_rows(min_row=2, values_only=True):
#         nome, ra, prova1, prova2, atividade, media, situacao = linha
#         situacao = ''
#         if nome == nome_param:
#             if media >= 6:
#                 situacao = 'APROVADO'
#             elif media < 5:
#                 situacao = 'REPROVADO'
#             else:
#                 situacao = 'RECUPERAÇÃO'
#     return 'Aluno não encontrado'


@app.route('/historico')
def historico():
    workbook = load_workbook(ARQUIVO)
    ws_historico = workbook.active
    dados = list(
        ws_historico.iter_rows(
            min_row=2,
            max_col=7,
            values_only=True
        )
    )
    return render_template('historico.html', dados=dados)


if __name__ == '__main__':
    app.run(debug=True)