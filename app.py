from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

ARQUIVO = 'cardapio.xlsx'

if not os.path.exists(ARQUIVO):
    wb = Workbook()
    ws = wb.active
    ws.append([
        'Imagem', 'Produto', 'Ingredientes',
        'Preço de custo', 'Margem', 'Preço de venda'
    ])
    wb.save(ARQUIVO)


@app.route('/')
def index():
    """
    Renders the 'index.html' template for the application's home page.

    Returns:
        str: Rendered HTML content of the index page.
    """
    return render_template('index.html')


@app.route('/salvar', methods=["POST"])
def salvar():
    imagem = request.form['imagem']
    nome = request.form['nome']
    ingrediente = (request.form['ingrediente'])
    custo = float(request.form['custo'])
    margem = float(request.form['margem'])
    venda = round((custo + (custo * (margem/100))), 2)
    workbook = load_workbook(ARQUIVO)
    ws_salvar = workbook.active
    ws_salvar.append([imagem, nome, ingrediente, custo, margem, venda])
    workbook.save(ARQUIVO)

    return render_template(
        'resultado.html',
        imagem=imagem,
        nome=nome,
        ingrediente=ingrediente,
        custo=custo,
        margem=margem,
        venda=venda
    )


# '@app.route('/analisar')
# def analisar():
#     nome_param = request.args.get('nome')
#     wb = load_workbook(ARQUIVO)
#     ws = wb.active
#     for linha in ws.iter_rows(min_row=2, values_only=True):
#         nome, ra, prova1, prova2, atividade, media, situacao = linha
#         situacao = ''
#         if nome == nome_param:
#             if media >= 6:
#                 situacao = 'APROVADO'
#             elif media < 5:
#                 situacao = 'REPROVADO'
#             else:
#                 situacao = 'RECUPERAÇÃO'
#     return 'Aluno não encontrado'


@app.route('/historico')
def historico():
    workbook = load_workbook(ARQUIVO)
    ws_historico = workbook.active
    dados = list(
        ws_historico.iter_rows(
            min_row=2, values_only=True)
            )
    return render_template('historico.html', dados=dados)


if __name__ == '__main__':
    app.run(debug=True)
# Este código é uma aplicação simples do Flask que gerencia um menu usando um
# arquivo do excel.